'''
Howl
'''
#coding=utf-8
#from openpyxl import Workbook
from openpyxl import load_workbook
import time

        
class OBJofXL(object):
    
    path = ""
    sheet_name = ""
    Customer_data = {u'備註': "", u'姓名': "", u'生日': "", u'根數': "", u'電話': "",\
                     u'睫毛次數': "", u'睫毛':True , u'除毛次數': "", u'入會日': "", u'最近消費日': "", u'上次重接日': "",\
                     u'預繳餘額': "", 'Line': ""}
    inwb = []
    workseet = []
    New_PSN = 0
    #Check Ann
    def __init__(self, arg1, arg2):
        
        self.path = arg1
        self.sheet_name = arg2 
        self.inwb = load_workbook(self.path)
        self.Check_workpage_sheetname(self.sheet_name)   
        self.workseet = self.inwb[self.sheet_name]
            
    def Check_workpage_sheetname(self, Fname):
        
        """
        Check if it it is for Fname file used Fname is the xlsx file name that you want to know
        """
        
        for self.sheetName in self.inwb.get_sheet_names():
            if self.sheetName == Fname:
                break
            else:
                continue
        else:
        #create work sheet
            self.inwb.create_sheet(Fname, 0)      
        self.inwb.save(self.path)

    def Save_Data(self, pos, data):
        '''
        Save your data into cell
        Arg 1 is the position of the data
        Arg 2 is the data value
        '''
    
    #set posision
        self.workseet[pos] = data
        self.inwb.save(self.path)

    def Save_Cell(self, row_update, col_update, data):
    
        '''
        Save your data into cell
        Arg 1 is row of data
        Arg 2 is colume of data
        Arg 3 is data
        '''
    
    #set posision
        self.workseet.cell(row = row_update, column = col_update).value = data
        self.inwb.save(self.path)
    
    def Read_Data(self, pos):
        
        '''Save your data in position'''
    
        return self.workseet[pos].value

    def Read_Cell(self, row_update, col_update):
        
        '''Save your data in position'''
    
        return self.workseet.cell(row = row_update, column = col_update)  
  
    def Add_new_member_data(self, member_data):
        '''
        looking for the last row and add customer data
        '''
        self.New_PSN = self.workseet.max_row + 1
        self.Save_Data('A'+str(self.New_PSN), self.workseet.max_row)
        self.Save_Data('B'+str(self.New_PSN), member_data[0])
        self.Save_Data('C'+str(self.New_PSN), member_data[1])
        self.Save_Data('D'+str(self.New_PSN), member_data[2])
        self.Save_Data('E'+str(self.New_PSN), member_data[3])#根數
        self.Save_Data('F'+str(self.New_PSN), member_data[4])
        #睫毛
        if member_data[5] == u"睫毛": 
            self.Save_Data('G'+str(self.New_PSN), 0)
        else:
            self.Save_Data('G'+str(self.New_PSN), 1)  
        #除毛         
        if member_data[6] == True:
            self.Save_Data('H'+str(self.New_PSN), 1)
        else:
            self.Save_Data('H'+str(self.New_PSN), 0)
        #入會日
        self.Save_Data('I'+str(self.New_PSN), time.strftime("%Y/%m/%d"))
        #上次消費日 
        self.Save_Data('J'+str(self.New_PSN), time.strftime("%Y/%m/%d")) 
        #重接
        if member_data[7] == True:
            self.Save_Data('K'+str(self.New_PSN), time.strftime("%Y/%m/%d"))  
        else:
            self.Save_Data('K'+str(self.New_PSN), 0)   
            
        self.Save_Data('L'+str(self.New_PSN), member_data[8]) # 預繳
        
        self.Save_Data('M'+str(self.New_PSN), member_data[9]) #line

    def Serch_member(self, compare_m, typeofmember):
        
        ''' 
        Get your member from xlsx
        Arg 1 is the vaule that you know
        Arg 2 is the type of this value
        '''
    
        for row in range(1, self.workseet.max_row + 1):
            if typeofmember == u'姓名':
                member = self.Read_Data('C'+str(row))
                if member == compare_m:
                    break
            if typeofmember == u'電話':
                member = self.Read_Data('F'+str(row))
                if member == compare_m:
                    break
        else:
            return False
        return row


    def Get_Member(self, row):
        
        ''' Get member data by row'''
 
        self.Customer_data[u'備註'] = self.Read_Data('B'+str(row))
        self.Customer_data[u'姓名'] = self.Read_Data('C'+str(row))
        self.Customer_data[u'生日'] = self.Read_Data('D'+str(row))
        self.Customer_data[u'根數'] = self.Read_Data('E'+str(row))
        self.Customer_data[u'電話'] = self.Read_Data('F'+str(row))
        self.Customer_data[u'睫毛次數'] = self.Read_Data('G'+str(row))
        self.Customer_data[u'除毛次數'] = self.Read_Data('H'+str(row))
        self.Customer_data[u'入會日'] = self.Read_Data('I'+str(row))
        self.Customer_data[u'最近消費日'] = self.Read_Data('J'+str(row))
        self.Customer_data[u'上次重接日'] = self.Read_Data('K'+str(row))
        self.Customer_data[u'預繳餘額'] = self.Read_Data('L'+str(row))
        self.Customer_data[u'Line'] = self.Read_Data('M'+str(row))
        #set to '' when None
        for key,value in self.Customer_data.items():
                if value == None :
                    self.Customer_data[key] = ''
        
    def Update_Member(self, row, member_data):
        '''
        
        '''
        print("Update start")
        self.Get_Member(row)
        self.Save_Data('B'+str(row), member_data[0])
        self.Save_Data('C'+str(row), member_data[1])
        self.Save_Data('D'+str(row), member_data[2])
        self.Save_Data('E'+str(row), member_data[3])
        self.Save_Data('F'+str(row), member_data[4])
        #睫毛
        if member_data[5] != u"睫毛": 
            self.Save_Data('G'+str(row), self.Customer_data[u'睫毛次數'] + 1)  
        #除毛         
        if member_data[6] == True:
            if self.Customer_data[u'除毛次數'] == None:
                self.Save_Data('H'+str(row), 1)
            else:
                self.Save_Data('H'+str(row), int(self.Customer_data[u'除毛次數']) + 1)
                
        #上次消費日 
        if member_data[7]!= "":# or member_data[7] != None:
        # some one add date
            self.Save_Data('J'+str(row), member_data[7])
        else:
        # default save current date
            self.Save_Data('J'+str(row), time.strftime("%Y/%m/%d")) 
        #重接
        if member_data[10] == True:
            if member_data[7] or member_data[7] != None:
                # some one add date
                self.Save_Data('K'+str(row), member_data[7])
            else:
                # default save current date    
                self.Save_Data('K'+str(row), time.strftime("%Y/%m/%d"))

        self.Save_Data('L'+str(row), member_data[8])
        
        self.Save_Data('M'+str(row), member_data[9]) #line
        print("Update it")
        
class OBJofXL2(OBJofXL):
    path = ""
    sheet_name = ""
    Customer_data = {'HEAD_PSN': "", u'日期': "", u'時間': "", u'翹度': "", u'根數': "",\
                     u'粗度': "", u'款式': "", u'長度': "", u'下睫毛': "", u'內容說明': "",\
                     u'消費金額': "", u'餘額': "", u'備註': ""}
    inwb = []
    workseet = []
    
    def __init__(self, arg1, arg2):
        self.path = arg1
        self.sheet_name = arg2 
        self.inwb = load_workbook(self.path)
        self.Check_workpage_sheetname(self.sheet_name)   
        self.workseet = self.inwb[self.sheet_name]
        
    def Check_workpage_sheetname(self, Fname):
        
        """
        Check if it it is for Fname file used Fname is the xlsx file name that you want to know
        """
        
        for self.sheetName in self.inwb.get_sheet_names():
            if self.sheetName == Fname:
                break
            else:
                continue
        else:
        #create work sheet
            self.inwb.create_sheet(Fname, 0)      
        self.inwb.save(self.path)
        
    def Get_Member(self, PSN):
        
        ''' Get member data by row'''
        
        row = self.Serch_member(PSN)
        if row != False:
            self.Customer_data['HEAD_PSN'] = self.Read_Data('B'+str(row))
            self.Customer_data[u'日期'] = self.Read_Data('C'+str(row))
            self.Customer_data[u'時間'] = self.Read_Data('D'+str(row))
            self.Customer_data[u'翹度'] = self.Read_Data('E'+str(row))
            self.Customer_data[u'根數'] = self.Read_Data('F'+str(row))
            self.Customer_data[u'粗度'] = self.Read_Data('G'+str(row))
            self.Customer_data[u'款式'] = self.Read_Data('H'+str(row))
            self.Customer_data[u'長度'] = self.Read_Data('I'+str(row))
            self.Customer_data[u'下睫毛'] = self.Read_Data('J'+str(row))
            self.Customer_data[u'內容說明'] = self.Read_Data('K'+str(row))
            self.Customer_data[u'儲值'] = self.Read_Data('L'+str(row))
            self.Customer_data[u'消費金額'] = self.Read_Data('M'+str(row))
            self.Customer_data[u'餘額'] = self.Read_Data('N'+str(row))
            self.Customer_data[u'備註'] = self.Read_Data('O'+str(row)) 
            #set to '' when None 
            for key,value in self.Customer_data.items():
                if value == None :
                    self.Customer_data[key] = ''
            return True
        else:
            return False          
        
    def Add_new_member_data(self, member_data):
        '''
        looking for the last row and add customer data
        '''
        print("Save start")
        new_row = self.workseet.max_row +1 
        self.Save_Data('A'+str(new_row), self.workseet.max_row)#flow
        self.Save_Data('B'+str(new_row), member_data[0]) #PSN
        if member_data[12]:
            self.Save_Data('C'+str(new_row), member_data[12])
        else:
            self.Save_Data('C'+str(new_row), time.strftime("%Y/%m/%d")) #date
            self.Save_Data('D'+str(new_row), time.strftime("%I:%M %p")) # time
        
        self.Save_Data('E'+str(new_row), member_data[1]) #翹度
        self.Save_Data('F'+str(new_row), member_data[2]) #根數
        self.Save_Data('G'+str(new_row), member_data[3]) #粗度
        self.Save_Data('H'+str(new_row), member_data[4]) #款式
        self.Save_Data('I'+str(new_row), member_data[5]) #長度
        if member_data[6] == True:
            self.Save_Data('J'+str(new_row), 'Y') #下睫毛
        else:
            self.Save_Data('J'+str(new_row), 'N') #下睫毛
        
        self.Save_Data('K'+str(new_row), member_data[7]) #內容
        #儲值
        self.Save_Data('L'+str(new_row), member_data[8])
        #消費金額
        self.Save_Data('M'+str(new_row), member_data[9])
        #餘額
        self.Save_Data('N'+str(new_row), member_data[10])
        self.Save_Data('O'+str(new_row), member_data[11])#備註
        
    def Serch_member(self, row):
        
        ''' 
        Get your member data from xlsx and return the row of last one 
        '''
        
        last_one = 0
        for serch_row in range(1, self.workseet.max_row + 1):
            member = self.Read_Data('B'+str(serch_row))
            if member == row:
                last_one = serch_row
                continue
        else:
            if last_one == 0:
                return False
        return last_one

'''
if __name__ == '__main__':
    DTL_path_in_use = './Ann_DTL.xlsx'
    DTL_page_in_use = 'CUST_DTL'       
    Setup_Data_Detail = OBJofXL2(DTL_path_in_use, DTL_page_in_use)
    Setup_Data_Detail.Check_workpage_sheetname(DTL_page_in_use)
    row = Setup_Data_Detail.Get_Member(5)
    print(Setup_Data_Detail.Customer_data.values())'''

    